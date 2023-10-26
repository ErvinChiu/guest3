# -*-coding:utf-8-*-
# @Time    :2023/10/259:31
# @Author  :Ervin Chiu
# @Email   :ErvinChiu@outlook.com
# @File    :do_excel.py
# @Software:PyCharm


from openpyxl import load_workbook


class DoExcel:

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def do_excel(self):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        test_data = []

        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data["id"] = sheet.cell(i, 1).value
            sub_data["title"] = sheet.cell(i, 2).value
            sub_data["method"] = sheet.cell(i, 3).value
            sub_data["url"] = sheet.cell(i, 4).value
            sub_data["params"] = sheet.cell(i, 5).value
            test_data.append(sub_data)

        return test_data

    def write_back(self, row, colum, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(row, colum).value = value
        wb.save(self.file_path)


if __name__ == '__main__':
    test_data = DoExcel(r"C:\Users\ErvinChiu\Desktop\data.xlsx", "Sheet1").do_excel()

    for item in test_data:
        print(item)
    test=DoExcel(r"C:\Users\ErvinChiu\Desktop\data.xlsx", "Sheet1")
    test.write_back(3,8,"pass")